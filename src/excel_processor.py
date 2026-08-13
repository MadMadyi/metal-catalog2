"""Excel processor for the Metal Catalog project.

This module connects Excel files with the metal catalogue pipeline:

Excel row
    -> metal_parser.py
    -> gost_finder.py
    -> metal_catalog.py
    -> result columns in Excel

Designed for .xlsx files.
"""

from __future__ import annotations

import argparse
from dataclasses import asdict, dataclass
from pathlib import Path
import re
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Font

from metal_catalog import analyze_metal


# Common header names used in working Excel files.
NAME_COLUMN_ALIASES = (
    "наименование",
    "наименование материала",
    "наименование тмц",
    "материал",
    "название",
    "номенклатура",
    "позиция",
    "полное наименование",
)

OUTPUT_HEADERS = {
    "metal_type": "Тип металлопроката",
    "size": "Размер",
    "gost": "Актуальный ГОСТ",
    "status": "Статус проверки",
    "comment": "Комментарий",
}

STATUS_MAP = {
    "high": "OK",
    "medium": "Проверить",
    "needs_review": "Требует проверки",
}


@dataclass(frozen=True)
class ExcelProcessSummary:
    input_path: str
    output_path: str
    sheet_name: str
    name_column: str
    processed_rows: int
    skipped_empty_rows: int
    high_confidence: int
    medium_confidence: int
    needs_review: int

    def to_dict(self) -> dict:
        return asdict(self)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text


def _column_letter(index: int) -> str:
    """Convert a 1-based column number to an Excel column letter."""
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _find_name_column(
    worksheet,
    *,
    header_row: int,
    name_header: Optional[str] = None,
) -> int:
    """Find the source column containing material descriptions."""
    headers = {
        column: _normalize_header(worksheet.cell(header_row, column).value)
        for column in range(1, worksheet.max_column + 1)
    }

    if name_header:
        requested = _normalize_header(name_header)
        for column, header in headers.items():
            if header == requested:
                return column
        raise ValueError(
            f'Столбец "{name_header}" не найден в строке заголовков {header_row}.'
        )

    # First try exact matches.
    for alias in NAME_COLUMN_ALIASES:
        for column, header in headers.items():
            if header == alias:
                return column

    # Then allow headers such as "Наименование / размер".
    for alias in NAME_COLUMN_ALIASES:
        for column, header in headers.items():
            if alias in header:
                return column

    available = [
        str(worksheet.cell(header_row, column).value)
        for column in range(1, worksheet.max_column + 1)
        if worksheet.cell(header_row, column).value not in (None, "")
    ]

    raise ValueError(
        "Не удалось автоматически найти столбец с наименованием материала. "
        f"Найдены заголовки: {available}. "
        "Передайте точное имя через --name-column."
    )


def _find_or_create_output_columns(worksheet, *, header_row: int) -> dict[str, int]:
    """Reuse existing result columns when possible, otherwise append them."""
    normalized_to_column = {
        _normalize_header(worksheet.cell(header_row, column).value): column
        for column in range(1, worksheet.max_column + 1)
    }

    result: dict[str, int] = {}
    next_column = worksheet.max_column + 1

    for key, header in OUTPUT_HEADERS.items():
        normalized = _normalize_header(header)

        if normalized in normalized_to_column:
            result[key] = normalized_to_column[normalized]
            continue

        column = next_column
        next_column += 1

        cell = worksheet.cell(header_row, column)
        cell.value = header
        cell.font = Font(bold=True)

        result[key] = column
        normalized_to_column[normalized] = column

    return result


def _build_comment(result) -> str:
    parts = [result.reason]

    if result.gost_candidates:
        parts.append("Варианты: " + ", ".join(result.gost_candidates))

    if result.future_replacement:
        future = f"Будущая замена: {result.future_replacement}"
        if result.future_effective_from:
            future += f" с {result.future_effective_from}"
        parts.append(future)

    return " | ".join(part for part in parts if part)


def _write_if_allowed(cell, value, *, overwrite: bool) -> None:
    if overwrite or cell.value in (None, ""):
        cell.value = value


def process_workbook(
    input_path: str | Path,
    output_path: str | Path | None = None,
    *,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    name_header: Optional[str] = None,
    overwrite: bool = False,
) -> ExcelProcessSummary:
    """Process one Excel workbook.

    Parameters
    ----------
    input_path:
        Source .xlsx file.
    output_path:
        Destination .xlsx file. If omitted, creates "<name>_processed.xlsx".
    sheet_name:
        Sheet to process. If omitted, uses the active sheet.
    header_row:
        1-based row number containing column headers.
    name_header:
        Exact header name for the material-description column.
        If omitted, the module tries to detect it automatically.
    overwrite:
        If False, existing non-empty result cells are preserved.
    """
    input_path = Path(input_path)

    if input_path.suffix.lower() != ".xlsx":
        raise ValueError("Сейчас поддерживаются только файлы .xlsx.")

    if not input_path.exists():
        raise FileNotFoundError(f"Файл не найден: {input_path}")

    if header_row < 1:
        raise ValueError("header_row должен быть не меньше 1.")

    if output_path is None:
        output_path = input_path.with_name(
            f"{input_path.stem}_processed{input_path.suffix}"
        )
    output_path = Path(output_path)

    workbook = load_workbook(input_path)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f'Лист "{sheet_name}" не найден. Доступные листы: {workbook.sheetnames}'
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    name_column = _find_name_column(
        worksheet,
        header_row=header_row,
        name_header=name_header,
    )

    output_columns = _find_or_create_output_columns(
        worksheet,
        header_row=header_row,
    )

    processed_rows = 0
    skipped_empty_rows = 0
    confidence_counts = {
        "high": 0,
        "medium": 0,
        "needs_review": 0,
    }

    for row in range(header_row + 1, worksheet.max_row + 1):
        raw_value = worksheet.cell(row, name_column).value

        if raw_value is None or not str(raw_value).strip():
            skipped_empty_rows += 1
            continue

        result = analyze_metal(str(raw_value))

        values = {
            "metal_type": result.metal_type or "",
            "size": result.size_text or "",
            "gost": result.gost or "",
            "status": STATUS_MAP.get(result.confidence, result.confidence),
            "comment": _build_comment(result),
        }

        for key, value in values.items():
            _write_if_allowed(
                worksheet.cell(row, output_columns[key]),
                value,
                overwrite=overwrite,
            )

        processed_rows += 1
        if result.confidence in confidence_counts:
            confidence_counts[result.confidence] += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return ExcelProcessSummary(
        input_path=str(input_path),
        output_path=str(output_path),
        sheet_name=worksheet.title,
        name_column=_column_letter(name_column),
        processed_rows=processed_rows,
        skipped_empty_rows=skipped_empty_rows,
        high_confidence=confidence_counts["high"],
        medium_confidence=confidence_counts["medium"],
        needs_review=confidence_counts["needs_review"],
    )


def _build_cli() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Обработка Excel-файлов через Metal Catalog."
    )

    parser.add_argument("input", help="Исходный .xlsx файл.")
    parser.add_argument(
        "output",
        nargs="?",
        help="Выходной .xlsx файл. Необязательно.",
    )
    parser.add_argument(
        "--sheet",
        dest="sheet_name",
        help="Имя листа. По умолчанию используется активный лист.",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help="Номер строки заголовков. По умолчанию: 1.",
    )
    parser.add_argument(
        "--name-column",
        dest="name_header",
        help='Точное название столбца с материалом, например "Наименование".',
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Перезаписывать уже заполненные результирующие ячейки.",
    )

    return parser


def main() -> None:
    parser = _build_cli()
    args = parser.parse_args()

    summary = process_workbook(
        args.input,
        args.output,
        sheet_name=args.sheet_name,
        header_row=args.header_row,
        name_header=args.name_header,
        overwrite=args.overwrite,
    )

    print("Обработка завершена.")
    print(f"Лист: {summary.sheet_name}")
    print(f"Столбец наименования: {summary.name_column}")
    print(f"Обработано строк: {summary.processed_rows}")
    print(f"Высокая уверенность: {summary.high_confidence}")
    print(f"Нужно проверить: {summary.medium_confidence}")
    print(f"Требует проверки: {summary.needs_review}")
    print(f"Результат: {summary.output_path}")


if __name__ == "__main__":
    main()
